"""app/services/data_management/export_ops.py

Export operations mixin for data management service.
Handles exporting recipes to xlsx and generating templates.
"""

# -- Imports -------------------------------------------------------------------------------------
from io import BytesIO
from typing import Optional

from openpyxl import Workbook

from ...dtos.data_management_dtos import ExportFilterDTO
from ...models import Recipe

from .service import INGREDIENT_COLUMNS, RECIPE_COLUMNS


# -- Export Operations Mixin ---------------------------------------------------------------------
class ExportOperationsMixin:
    """Mixin providing export operations (xlsx export, template generation)."""

    # -- Export ----------------------------------------------------------------------------------
    def export_recipes_to_xlsx(
        self, filter_dto: Optional[ExportFilterDTO] = None
    ) -> bytes:
        """
        Export recipes to xlsx format.

        Args:
            filter_dto: Optional filters for which recipes to export.

        Returns:
            Bytes of the xlsx file.
        """
        # Build query
        query = self.session.query(Recipe)

        if filter_dto:
            if filter_dto.recipe_category:
                query = query.filter(Recipe.recipe_category == filter_dto.recipe_category)
            if filter_dto.meal_type:
                query = query.filter(Recipe.meal_type == filter_dto.meal_type)
            if filter_dto.favorites_only:
                query = query.filter(Recipe.is_favorite == True)

        recipes = query.all()

        # Create workbook
        workbook = Workbook()

        # Recipes sheet
        recipes_sheet = workbook.active
        recipes_sheet.title = "Recipes"
        recipes_sheet.append(RECIPE_COLUMNS)

        for recipe in recipes:
            recipes_sheet.append([
                recipe.recipe_name,
                recipe.recipe_category,
                recipe.meal_type,
                recipe.diet_pref,
                recipe.total_time,
                recipe.servings,
                recipe.directions,
                recipe.notes,
            ])

        # Ingredients sheet
        ingredients_sheet = workbook.create_sheet("Ingredients")
        ingredients_sheet.append(INGREDIENT_COLUMNS)

        for recipe in recipes:
            for ing in recipe.ingredients:
                ingredients_sheet.append([
                    recipe.recipe_name,
                    recipe.recipe_category,
                    ing.ingredient.ingredient_name,
                    ing.ingredient.ingredient_category,
                    ing.quantity,
                    ing.unit,
                ])

        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_template_xlsx(self) -> bytes:
        """
        Generate an empty xlsx template with correct headers.

        Returns:
            Bytes of the template xlsx file.
        """
        workbook = Workbook()

        # Recipes sheet
        recipes_sheet = workbook.active
        recipes_sheet.title = "Recipes"
        recipes_sheet.append(RECIPE_COLUMNS)

        # Add example row
        recipes_sheet.append([
            "Example Recipe",
            "Main Course",
            "Dinner",
            "",
            30,
            4,
            "1. Step one\n2. Step two",
            "Optional notes here",
        ])

        # Ingredients sheet
        ingredients_sheet = workbook.create_sheet("Ingredients")
        ingredients_sheet.append(INGREDIENT_COLUMNS)

        # Add example rows
        ingredients_sheet.append([
            "Example Recipe",
            "Main Course",
            "Chicken Breast",
            "Meat",
            2,
            "lbs",
        ])
        ingredients_sheet.append([
            "Example Recipe",
            "Main Course",
            "Olive Oil",
            "Pantry",
            2,
            "tbsp",
        ])

        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
