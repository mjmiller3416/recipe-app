"""app/core/services/data_management_service.py

Service for data management operations (import/export recipes via xlsx).
"""

# ── Imports ─────────────────────────────────────────────────────────────────────────────────────────────────
import os
import re
from datetime import datetime, timezone
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import cloudinary
import cloudinary.uploader
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func
from sqlalchemy.orm import Session

# Load environment variables for Cloudinary config
load_dotenv()

# Configure Cloudinary
cloudinary.config(
    cloud_name=os.getenv("CLOUDINARY_CLOUD_NAME"),
    api_key=os.getenv("CLOUDINARY_API_KEY"),
    api_secret=os.getenv("CLOUDINARY_API_SECRET"),
    secure=True
)

from ..dtos.data_management_dtos import (
    BackupDataDTO,
    DuplicateAction,
    DuplicateRecipeDTO,
    DuplicateResolutionDTO,
    ExportFilterDTO,
    FullBackupDTO,
    ImportPreviewDTO,
    ImportResultDTO,
    IngredientBackupDTO,
    MealBackupDTO,
    PlannerEntryBackupDTO,
    RecipeBackupDTO,
    RecipeHistoryBackupDTO,
    RecipeImportRowDTO,
    RecipeIngredientBackupDTO,
    RestorePreviewDTO,
    RestoreResultDTO,
    ShoppingItemBackupDTO,
    ValidationErrorDTO,
)
from ..dtos.recipe_dtos import RecipeCreateDTO, RecipeIngredientDTO, RecipeUpdateDTO
from ..models import (
    Ingredient,
    Meal,
    PlannerEntry,
    Recipe,
    RecipeHistory,
    RecipeIngredient,
    ShoppingItem,
    ShoppingItemContribution,
)
from ..repositories.ingredient_repo import IngredientRepo
from ..repositories.recipe_repo import RecipeRepo


# ── Constants ───────────────────────────────────────────────────────────────────────────────────────────────
RECIPE_COLUMNS = [
    "recipe_name",
    "recipe_category",
    "meal_type",
    "diet_pref",
    "total_time",
    "servings",
    "directions",
    "notes",
]

INGREDIENT_COLUMNS = [
    "recipe_name",
    "recipe_category",
    "ingredient_name",
    "ingredient_category",
    "quantity",
    "unit",
]


# ── Data Management Service ─────────────────────────────────────────────────────────────────────────────────
class DataManagementService:
    """Service layer for importing and exporting recipe data."""

    def __init__(self, session: Session, user_id: int = None):
        """Initialize the service with a database session and optional user ID.

        Args:
            session: SQLAlchemy database session
            user_id: ID of the current user for multi-tenant operations (required for write operations)
        """
        self.session = session
        self.user_id = user_id
        # Ingredient repo requires user_id for user-scoped operations
        # For read-only operations (template generation), user_id can be None
        if user_id:
            self.ingredient_repo = IngredientRepo(session, user_id)
            self.recipe_repo = RecipeRepo(session, self.ingredient_repo, user_id)
        else:
            self.ingredient_repo = None
            self.recipe_repo = RecipeRepo(session, user_id=None)

    # ── Parsing ─────────────────────────────────────────────────────────────────────────────────────────────
    def parse_xlsx(
        self, file_content: bytes
    ) -> Tuple[List[RecipeImportRowDTO], List[ValidationErrorDTO]]:
        """
        Parse xlsx file content into recipe DTOs.

        Args:
            file_content: Raw bytes of the xlsx file.

        Returns:
            Tuple of (parsed recipes, validation errors).
        """
        recipes: List[RecipeImportRowDTO] = []
        errors: List[ValidationErrorDTO] = []

        try:
            workbook = load_workbook(filename=BytesIO(file_content), read_only=True)
        except Exception as e:
            errors.append(
                ValidationErrorDTO(
                    row_number=0, field="file", message=f"Invalid xlsx file: {str(e)}"
                )
            )
            return recipes, errors

        # Parse Recipes sheet
        if "Recipes" not in workbook.sheetnames:
            errors.append(
                ValidationErrorDTO(
                    row_number=0,
                    field="sheet",
                    message="Missing required 'Recipes' sheet",
                )
            )
            return recipes, errors

        recipes_sheet = workbook["Recipes"]
        recipe_rows, recipe_errors = self._parse_recipes_sheet(recipes_sheet)
        errors.extend(recipe_errors)

        # Parse Ingredients sheet (optional)
        ingredients_by_recipe: Dict[Tuple[str, str], List[RecipeIngredientDTO]] = {}
        if "Ingredients" in workbook.sheetnames:
            ingredients_sheet = workbook["Ingredients"]
            ingredients_by_recipe, ing_errors = self._parse_ingredients_sheet(
                ingredients_sheet
            )
            errors.extend(ing_errors)

        # Combine recipes with their ingredients
        for recipe_data in recipe_rows:
            key = (
                recipe_data["recipe_name"].lower(),
                recipe_data["recipe_category"].lower(),
            )
            ingredients = ingredients_by_recipe.get(key, [])

            try:
                recipe = RecipeImportRowDTO(
                    recipe_name=recipe_data["recipe_name"],
                    recipe_category=recipe_data["recipe_category"],
                    meal_type=recipe_data.get("meal_type") or "Dinner",
                    diet_pref=recipe_data.get("diet_pref"),
                    total_time=recipe_data.get("total_time"),
                    servings=recipe_data.get("servings"),
                    directions=recipe_data.get("directions"),
                    notes=recipe_data.get("notes"),
                    ingredients=ingredients,
                )
                recipes.append(recipe)
            except Exception as e:
                errors.append(
                    ValidationErrorDTO(
                        row_number=recipe_data.get("_row", 0),
                        field="recipe",
                        message=str(e),
                    )
                )

        workbook.close()
        return recipes, errors

    def _parse_recipes_sheet(
        self, sheet: Worksheet
    ) -> Tuple[List[Dict], List[ValidationErrorDTO]]:
        """Parse the Recipes sheet into row dictionaries."""
        rows: List[Dict] = []
        errors: List[ValidationErrorDTO] = []

        # Get header row
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = [h.lower().strip() if h else "" for h in headers]

        # Validate required columns
        if "recipe_name" not in headers or "recipe_category" not in headers:
            errors.append(
                ValidationErrorDTO(
                    row_number=1,
                    field="headers",
                    message="Missing required columns: recipe_name, recipe_category",
                )
            )
            return rows, errors

        # Map column indices
        col_map = {h: i for i, h in enumerate(headers) if h}

        # Parse data rows
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            cells = [cell.value for cell in row]

            # Skip empty rows
            if not any(cells):
                continue

            recipe_name = self._get_cell_value(cells, col_map, "recipe_name")
            recipe_category = self._get_cell_value(cells, col_map, "recipe_category")

            # Validate required fields
            if not recipe_name:
                errors.append(
                    ValidationErrorDTO(
                        row_number=row_num,
                        field="recipe_name",
                        message="Recipe name is required",
                    )
                )
                continue

            if not recipe_category:
                errors.append(
                    ValidationErrorDTO(
                        row_number=row_num,
                        field="recipe_category",
                        message="Recipe category is required",
                    )
                )
                continue

            # Normalize category and meal_type to lowercase for frontend compatibility
            meal_type_raw = self._get_cell_value(cells, col_map, "meal_type")
            diet_pref_raw = self._get_cell_value(cells, col_map, "diet_pref")

            row_data = {
                "_row": row_num,
                "recipe_name": str(recipe_name).strip(),
                "recipe_category": str(recipe_category).strip().lower(),
                "meal_type": str(meal_type_raw).strip().lower() if meal_type_raw else None,
                "diet_pref": str(diet_pref_raw).strip().lower() if diet_pref_raw else None,
                "total_time": self._parse_int(
                    self._get_cell_value(cells, col_map, "total_time")
                ),
                "servings": self._parse_int(
                    self._get_cell_value(cells, col_map, "servings")
                ),
                "directions": self._get_cell_value(cells, col_map, "directions"),
                "notes": self._get_cell_value(cells, col_map, "notes"),
            }
            rows.append(row_data)

        return rows, errors

    def _parse_ingredients_sheet(
        self, sheet: Worksheet
    ) -> Tuple[Dict[Tuple[str, str], List[RecipeIngredientDTO]], List[ValidationErrorDTO]]:
        """Parse the Ingredients sheet into a dict keyed by (recipe_name, recipe_category)."""
        ingredients: Dict[Tuple[str, str], List[RecipeIngredientDTO]] = {}
        errors: List[ValidationErrorDTO] = []

        # Get header row
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = [h.lower().strip() if h else "" for h in headers]

        # Validate required columns
        required = ["recipe_name", "recipe_category", "ingredient_name", "ingredient_category"]
        missing = [col for col in required if col not in headers]
        if missing:
            errors.append(
                ValidationErrorDTO(
                    row_number=1,
                    field="headers",
                    message=f"Missing required columns: {', '.join(missing)}",
                )
            )
            return ingredients, errors

        col_map = {h: i for i, h in enumerate(headers) if h}

        # Parse data rows
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            cells = [cell.value for cell in row]

            # Skip empty rows
            if not any(cells):
                continue

            recipe_name = self._get_cell_value(cells, col_map, "recipe_name")
            recipe_category = self._get_cell_value(cells, col_map, "recipe_category")
            ingredient_name = self._get_cell_value(cells, col_map, "ingredient_name")
            ingredient_category = self._get_cell_value(cells, col_map, "ingredient_category")

            if not all([recipe_name, recipe_category, ingredient_name, ingredient_category]):
                errors.append(
                    ValidationErrorDTO(
                        row_number=row_num,
                        field="ingredient",
                        message="Missing required ingredient fields",
                    )
                )
                continue

            recipe_key = (str(recipe_name).lower().strip(), str(recipe_category).lower().strip())
            quantity = self._parse_float(self._get_cell_value(cells, col_map, "quantity"))
            unit = self._get_cell_value(cells, col_map, "unit")

            ing_name = str(ingredient_name).strip()
            ing_cat = str(ingredient_category).strip().lower()  # Normalize to lowercase
            ing_unit = str(unit).strip().lower() if unit else None  # Normalize to lowercase

            if recipe_key not in ingredients:
                ingredients[recipe_key] = {}

            # Deduplicate by ingredient name + category (combine quantities)
            ing_key = (ing_name.lower(), ing_cat.lower())
            if ing_key in ingredients[recipe_key]:
                # Combine quantities if same unit, otherwise keep first
                existing = ingredients[recipe_key][ing_key]
                if quantity and existing.quantity and existing.unit == ing_unit:
                    existing.quantity += quantity
            else:
                ingredients[recipe_key][ing_key] = RecipeIngredientDTO(
                    ingredient_name=ing_name,
                    ingredient_category=ing_cat,
                    quantity=quantity,
                    unit=ing_unit,
                )

        # Convert dict values to lists
        result: Dict[Tuple[str, str], List[RecipeIngredientDTO]] = {}
        for recipe_key, ing_dict in ingredients.items():
            result[recipe_key] = list(ing_dict.values())

        return result, errors

    def _get_cell_value(self, cells: List, col_map: Dict[str, int], column: str):
        """Get cell value by column name, returns None if column doesn't exist."""
        if column not in col_map:
            return None
        idx = col_map[column]
        if idx >= len(cells):
            return None
        value = cells[idx]
        if isinstance(value, str):
            value = value.strip()
            return value if value else None
        return value

    def _parse_int(self, value) -> Optional[int]:
        """Parse a value to int, returns None if invalid."""
        if value is None:
            return None
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None

    def _parse_float(self, value) -> Optional[float]:
        """Parse a value to float, returns None if invalid."""
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    # ── Preview ─────────────────────────────────────────────────────────────────────────────────────────────
    def get_import_preview(
        self, recipes: List[RecipeImportRowDTO]
    ) -> ImportPreviewDTO:
        """
        Check parsed recipes against database to identify duplicates.

        Args:
            recipes: List of parsed recipe DTOs.

        Returns:
            ImportPreviewDTO with counts and duplicate info.
        """
        duplicates: List[DuplicateRecipeDTO] = []
        new_count = 0

        for idx, recipe in enumerate(recipes, start=2):  # Row 2 is first data row
            existing = self._find_existing_recipe(
                recipe.recipe_name, recipe.recipe_category
            )
            if existing:
                duplicates.append(
                    DuplicateRecipeDTO(
                        recipe_name=recipe.recipe_name,
                        recipe_category=recipe.recipe_category,
                        existing_id=existing.id,
                        row_number=idx,
                    )
                )
            else:
                new_count += 1

        return ImportPreviewDTO(
            total_recipes=len(recipes),
            new_recipes=new_count,
            duplicate_recipes=duplicates,
            validation_errors=[],
        )

    def _find_existing_recipe(self, name: str, category: str) -> Optional[Recipe]:
        """Find existing recipe by name and category (case-insensitive)."""
        return (
            self.session.query(Recipe)
            .filter(
                func.lower(Recipe.recipe_name) == name.strip().lower(),
                func.lower(Recipe.recipe_category) == category.strip().lower(),
            )
            .first()
        )

    # ── Execute Import ──────────────────────────────────────────────────────────────────────────────────────
    def execute_import(
        self,
        recipes: List[RecipeImportRowDTO],
        resolutions: List[DuplicateResolutionDTO],
    ) -> ImportResultDTO:
        """
        Execute the import with user-specified duplicate resolutions.

        Args:
            recipes: List of parsed recipe DTOs.
            resolutions: User's choices for handling duplicates.

        Returns:
            ImportResultDTO with counts and any errors.
        """
        # Build resolution lookup
        resolution_map: Dict[Tuple[str, str], DuplicateResolutionDTO] = {}
        for res in resolutions:
            key = (res.recipe_name.lower(), res.recipe_category.lower())
            resolution_map[key] = res

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors: List[str] = []

        for recipe in recipes:
            key = (recipe.recipe_name.lower(), recipe.recipe_category.lower())
            existing = self._find_existing_recipe(
                recipe.recipe_name, recipe.recipe_category
            )

            try:
                if existing:
                    resolution = resolution_map.get(key)
                    if resolution is None or resolution.action == DuplicateAction.SKIP:
                        skipped_count += 1
                        continue
                    elif resolution.action == DuplicateAction.UPDATE:
                        self._update_existing_recipe(existing, recipe)
                        updated_count += 1
                    elif resolution.action == DuplicateAction.RENAME:
                        if not resolution.new_name:
                            errors.append(
                                f"No new name provided for '{recipe.recipe_name}'"
                            )
                            skipped_count += 1
                            continue
                        self._create_recipe(recipe, new_name=resolution.new_name)
                        created_count += 1
                else:
                    self._create_recipe(recipe)
                    created_count += 1

            except Exception as e:
                errors.append(f"Error importing '{recipe.recipe_name}': {str(e)}")
                # Rollback to clear the failed transaction so we can continue
                self.session.rollback()

        # Commit all changes
        try:
            self.session.commit()
        except Exception as e:
            self.session.rollback()
            return ImportResultDTO(
                success=False,
                created_count=0,
                updated_count=0,
                skipped_count=len(recipes),
                errors=[f"Database error: {str(e)}"],
            )

        return ImportResultDTO(
            success=len(errors) == 0,
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=skipped_count,
            errors=errors,
        )

    def _create_recipe(
        self, recipe: RecipeImportRowDTO, new_name: Optional[str] = None
    ) -> Recipe:
        """Create a new recipe from import data."""
        create_dto = RecipeCreateDTO(
            recipe_name=new_name or recipe.recipe_name,
            recipe_category=recipe.recipe_category,
            meal_type=recipe.meal_type,
            diet_pref=recipe.diet_pref,
            total_time=recipe.total_time,
            servings=recipe.servings,
            directions=recipe.directions,
            notes=recipe.notes,
            ingredients=recipe.ingredients,
        )
        return self.recipe_repo.persist_recipe_and_links(create_dto)

    def _update_existing_recipe(
        self, existing: Recipe, recipe: RecipeImportRowDTO
    ) -> Recipe:
        """Update an existing recipe with import data."""
        update_dto = RecipeUpdateDTO(
            meal_type=recipe.meal_type,
            diet_pref=recipe.diet_pref,
            total_time=recipe.total_time,
            servings=recipe.servings,
            directions=recipe.directions,
            notes=recipe.notes,
            ingredients=recipe.ingredients,
        )
        return self.recipe_repo.update_recipe(existing.id, update_dto)

    # ── Export ──────────────────────────────────────────────────────────────────────────────────────────────
    def export_recipes_to_xlsx(
        self, filter_dto: Optional[ExportFilterDTO] = None
    ) -> bytes:
        """
        Export recipes to xlsx format.

        Args:
            filter_dto: Optional filters for which recipes to export.

        Returns:
            Bytes of the xlsx file.
        """
        # Build query
        query = self.session.query(Recipe)

        if filter_dto:
            if filter_dto.recipe_category:
                query = query.filter(Recipe.recipe_category == filter_dto.recipe_category)
            if filter_dto.meal_type:
                query = query.filter(Recipe.meal_type == filter_dto.meal_type)
            if filter_dto.favorites_only:
                query = query.filter(Recipe.is_favorite == True)

        recipes = query.all()

        # Create workbook
        workbook = Workbook()

        # Recipes sheet
        recipes_sheet = workbook.active
        recipes_sheet.title = "Recipes"
        recipes_sheet.append(RECIPE_COLUMNS)

        for recipe in recipes:
            recipes_sheet.append([
                recipe.recipe_name,
                recipe.recipe_category,
                recipe.meal_type,
                recipe.diet_pref,
                recipe.total_time,
                recipe.servings,
                recipe.directions,
                recipe.notes,
            ])

        # Ingredients sheet
        ingredients_sheet = workbook.create_sheet("Ingredients")
        ingredients_sheet.append(INGREDIENT_COLUMNS)

        for recipe in recipes:
            for ing in recipe.ingredients:
                ingredients_sheet.append([
                    recipe.recipe_name,
                    recipe.recipe_category,
                    ing.ingredient.ingredient_name,
                    ing.ingredient.ingredient_category,
                    ing.quantity,
                    ing.unit,
                ])

        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_template_xlsx(self) -> bytes:
        """
        Generate an empty xlsx template with correct headers.

        Returns:
            Bytes of the template xlsx file.
        """
        workbook = Workbook()

        # Recipes sheet
        recipes_sheet = workbook.active
        recipes_sheet.title = "Recipes"
        recipes_sheet.append(RECIPE_COLUMNS)

        # Add example row
        recipes_sheet.append([
            "Example Recipe",
            "Main Course",
            "Dinner",
            "",
            30,
            4,
            "1. Step one\n2. Step two",
            "Optional notes here",
        ])

        # Ingredients sheet
        ingredients_sheet = workbook.create_sheet("Ingredients")
        ingredients_sheet.append(INGREDIENT_COLUMNS)

        # Add example rows
        ingredients_sheet.append([
            "Example Recipe",
            "Main Course",
            "Chicken Breast",
            "Meat",
            2,
            "lbs",
        ])
        ingredients_sheet.append([
            "Example Recipe",
            "Main Course",
            "Olive Oil",
            "Pantry",
            2,
            "tbsp",
        ])

        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    # ── Clear All Data ─────────────────────────────────────────────────────────────────────────────────────────
    def _extract_cloudinary_public_id(self, url: str) -> Optional[str]:
        """
        Extract the public_id from a Cloudinary URL.

        Cloudinary URLs look like:
        https://res.cloudinary.com/{cloud}/image/upload/v{version}/{public_id}.{ext}

        Returns the public_id without the file extension.
        """
        if not url:
            return None

        # Match the path after /upload/ and before the file extension
        # Example: .../upload/v1234567890/meal-genie/recipes/123/reference_123.jpg
        match = re.search(r"/upload/(?:v\d+/)?(.+)\.\w+$", url)
        if match:
            return match.group(1)
        return None

    def _delete_cloudinary_images(self, recipes: List[Recipe]) -> int:
        """
        Delete all Cloudinary images for the given recipes.

        Returns the count of successfully deleted images.
        """
        deleted_count = 0

        for recipe in recipes:
            for image_path in [recipe.reference_image_path, recipe.banner_image_path]:
                if image_path:
                    public_id = self._extract_cloudinary_public_id(image_path)
                    if public_id:
                        try:
                            result = cloudinary.uploader.destroy(public_id)
                            if result.get("result") == "ok":
                                deleted_count += 1
                        except Exception:
                            # Continue even if individual deletion fails
                            pass

        return deleted_count

    def clear_all_data(self) -> Dict[str, int]:
        """
        Delete all data from all tables, including Cloudinary images.

        Deletes Cloudinary images first, then tables in the correct order
        to respect foreign key constraints.

        Returns:
            Dict with counts of deleted records per table.
        """
        counts = {}

        # First, delete Cloudinary images before removing recipe records
        recipes_with_images = (
            self.session.query(Recipe)
            .filter(
                (Recipe.reference_image_path.isnot(None))
                | (Recipe.banner_image_path.isnot(None))
            )
            .all()
        )
        counts["cloudinary_images"] = self._delete_cloudinary_images(recipes_with_images)

        # Delete in order to respect foreign key constraints
        # Shopping contributions depend on ShoppingItem
        counts["shopping_contributions"] = self.session.query(ShoppingItemContribution).delete()

        # Shopping items
        counts["shopping_items"] = self.session.query(ShoppingItem).delete()

        # Planner entries depend on Recipe
        counts["planner_entries"] = self.session.query(PlannerEntry).delete()

        # Recipe ingredients depend on Recipe and Ingredient
        counts["recipe_ingredients"] = self.session.query(RecipeIngredient).delete()

        # Recipe history depends on Recipe
        counts["recipe_history"] = self.session.query(RecipeHistory).delete()

        # Recipes
        counts["recipes"] = self.session.query(Recipe).delete()

        # Meals
        counts["meals"] = self.session.query(Meal).delete()

        # Ingredients (can be deleted after recipe_ingredients)
        counts["ingredients"] = self.session.query(Ingredient).delete()

        self.session.commit()

        return counts

    # ── Full Backup/Restore ────────────────────────────────────────────────────────────────────────────────────────
    def export_full_backup(self) -> FullBackupDTO:
        """
        Export all database data as a FullBackupDTO.

        Settings are not included here (they come from frontend localStorage).

        Returns:
            FullBackupDTO with all database data.
        """
        # Query all tables
        ingredients = self.session.query(Ingredient).all()
        recipes = self.session.query(Recipe).all()
        recipe_ingredients = self.session.query(RecipeIngredient).all()
        recipe_history = self.session.query(RecipeHistory).all()
        meals = self.session.query(Meal).all()
        planner_entries = self.session.query(PlannerEntry).all()
        shopping_items = self.session.query(ShoppingItem).all()

        # Convert to DTOs
        return FullBackupDTO(
            created_at=datetime.now(timezone.utc),
            data=BackupDataDTO(
                ingredients=[
                    IngredientBackupDTO(
                        id=ing.id,
                        ingredient_name=ing.ingredient_name,
                        ingredient_category=ing.ingredient_category,
                    )
                    for ing in ingredients
                ],
                recipes=[
                    RecipeBackupDTO(
                        id=r.id,
                        recipe_name=r.recipe_name,
                        recipe_category=r.recipe_category,
                        meal_type=r.meal_type,
                        diet_pref=r.diet_pref,
                        total_time=r.total_time,
                        servings=r.servings,
                        directions=r.directions,
                        notes=r.notes,
                        reference_image_path=r.reference_image_path,
                        banner_image_path=r.banner_image_path,
                        created_at=r.created_at,
                        is_favorite=r.is_favorite,
                    )
                    for r in recipes
                ],
                recipe_ingredients=[
                    RecipeIngredientBackupDTO(
                        recipe_id=ri.recipe_id,
                        ingredient_id=ri.ingredient_id,
                        quantity=ri.quantity,
                        unit=ri.unit,
                    )
                    for ri in recipe_ingredients
                ],
                recipe_history=[
                    RecipeHistoryBackupDTO(
                        id=rh.id,
                        recipe_id=rh.recipe_id,
                        cooked_at=rh.cooked_at,
                    )
                    for rh in recipe_history
                ],
                meals=[
                    MealBackupDTO(
                        id=m.id,
                        meal_name=m.meal_name,
                        main_recipe_id=m.main_recipe_id,
                        side_recipe_ids=m.side_recipe_ids,
                        tags=m.tags,
                        is_saved=m.is_saved,
                        created_at=m.created_at,
                    )
                    for m in meals
                ],
                planner_entries=[
                    PlannerEntryBackupDTO(
                        id=pe.id,
                        meal_id=pe.meal_id,
                        position=pe.position,
                        is_completed=pe.is_completed,
                        completed_at=pe.completed_at,
                        scheduled_date=pe.scheduled_date,
                        shopping_mode=pe.shopping_mode,
                        is_cleared=pe.is_cleared,
                    )
                    for pe in planner_entries
                ],
                shopping_items=[
                    ShoppingItemBackupDTO(
                        id=si.id,
                        ingredient_name=si.ingredient_name,
                        quantity=si.quantity,
                        unit=si.unit,
                        category=si.category,
                        source=si.source,
                        have=si.have,
                        flagged=si.flagged,
                        state_key=si.aggregation_key,  # Use aggregation_key for backwards compat
                        recipe_sources=[],  # Contributions are rebuilt on restore via sync
                    )
                    for si in shopping_items
                ],
                # shopping_states removed - state now lives on ShoppingItem directly
                shopping_states=[],
            ),
        )

    def preview_restore(self, backup: FullBackupDTO) -> RestorePreviewDTO:
        """
        Preview what will be restored without making changes.

        Args:
            backup: The backup data to preview.

        Returns:
            RestorePreviewDTO with counts and warnings.
        """
        counts = {
            "ingredients": len(backup.data.ingredients),
            "recipes": len(backup.data.recipes),
            "recipe_ingredients": len(backup.data.recipe_ingredients),
            "recipe_history": len(backup.data.recipe_history),
            "meals": len(backup.data.meals),
            "planner_entries": len(backup.data.planner_entries),
            "shopping_items": len(backup.data.shopping_items),
        }

        warnings: List[str] = []

        # Check version compatibility
        if backup.version != "1.0.0":
            warnings.append(f"Backup version {backup.version} may not be fully compatible")

        # Check for existing data
        existing_recipes = self.session.query(Recipe).count()
        if existing_recipes > 0:
            warnings.append(f"Existing data ({existing_recipes} recipes) will be deleted before restore")

        return RestorePreviewDTO(
            backup_version=backup.version,
            backup_created_at=backup.created_at,
            counts=counts,
            has_settings=backup.settings is not None,
            warnings=warnings,
        )

    def execute_restore(
        self, backup: FullBackupDTO, clear_existing: bool = True
    ) -> RestoreResultDTO:
        """
        Execute a full restore from backup.

        Restores data in dependency order to maintain referential integrity:
        1. Ingredients (no dependencies)
        2. Recipes (no dependencies)
        3. Recipe Ingredients (depends on recipes + ingredients)
        4. Recipe History (depends on recipes)
        5. Meals (depends on recipes)
        6. Planner Entries (depends on meals)
        7. Shopping States (no dependencies)
        8. Shopping Items (no dependencies)

        Args:
            backup: The backup data to restore.
            clear_existing: If True, clear all existing data first.

        Returns:
            RestoreResultDTO with counts and any errors.
        """
        errors: List[str] = []
        restored_counts: Dict[str, int] = {}

        try:
            if clear_existing:
                self.clear_all_data()

            # Build ID mappings as we restore (old_id -> new_id)
            ingredient_id_map: Dict[int, int] = {}
            recipe_id_map: Dict[int, int] = {}
            meal_id_map: Dict[int, int] = {}

            # 1. Restore ingredients
            for ing_dto in backup.data.ingredients:
                new_ing = Ingredient(
                    ingredient_name=ing_dto.ingredient_name,
                    ingredient_category=ing_dto.ingredient_category,
                )
                self.session.add(new_ing)
                self.session.flush()
                ingredient_id_map[ing_dto.id] = new_ing.id
            restored_counts["ingredients"] = len(backup.data.ingredients)

            # 2. Restore recipes
            for recipe_dto in backup.data.recipes:
                new_recipe = Recipe(
                    recipe_name=recipe_dto.recipe_name,
                    recipe_category=recipe_dto.recipe_category,
                    meal_type=recipe_dto.meal_type,
                    diet_pref=recipe_dto.diet_pref,
                    total_time=recipe_dto.total_time,
                    servings=recipe_dto.servings,
                    directions=recipe_dto.directions,
                    notes=recipe_dto.notes,
                    reference_image_path=recipe_dto.reference_image_path,
                    banner_image_path=recipe_dto.banner_image_path,
                    created_at=recipe_dto.created_at,
                    is_favorite=recipe_dto.is_favorite,
                )
                self.session.add(new_recipe)
                self.session.flush()
                recipe_id_map[recipe_dto.id] = new_recipe.id
            restored_counts["recipes"] = len(backup.data.recipes)

            # 3. Restore recipe ingredients (using mapped IDs)
            for ri_dto in backup.data.recipe_ingredients:
                new_recipe_id = recipe_id_map.get(ri_dto.recipe_id)
                new_ingredient_id = ingredient_id_map.get(ri_dto.ingredient_id)
                if new_recipe_id and new_ingredient_id:
                    new_ri = RecipeIngredient(
                        recipe_id=new_recipe_id,
                        ingredient_id=new_ingredient_id,
                        quantity=ri_dto.quantity,
                        unit=ri_dto.unit,
                    )
                    self.session.add(new_ri)
                else:
                    errors.append(f"Skipped recipe ingredient: missing recipe or ingredient reference")
            restored_counts["recipe_ingredients"] = len(backup.data.recipe_ingredients)

            # 4. Restore recipe history (using mapped recipe IDs)
            for rh_dto in backup.data.recipe_history:
                new_recipe_id = recipe_id_map.get(rh_dto.recipe_id)
                if new_recipe_id:
                    new_rh = RecipeHistory(
                        recipe_id=new_recipe_id,
                        cooked_at=rh_dto.cooked_at,
                    )
                    self.session.add(new_rh)
                else:
                    errors.append(f"Skipped recipe history: missing recipe reference")
            restored_counts["recipe_history"] = len(backup.data.recipe_history)

            # 5. Restore meals (using mapped recipe IDs)
            for meal_dto in backup.data.meals:
                new_main_recipe_id = recipe_id_map.get(meal_dto.main_recipe_id)
                if new_main_recipe_id:
                    new_side_ids = [
                        recipe_id_map.get(sid)
                        for sid in meal_dto.side_recipe_ids
                        if recipe_id_map.get(sid)
                    ]
                    new_meal = Meal(
                        meal_name=meal_dto.meal_name,
                        main_recipe_id=new_main_recipe_id,
                        is_saved=meal_dto.is_saved,
                        created_at=meal_dto.created_at,
                    )
                    new_meal.side_recipe_ids = new_side_ids
                    new_meal.tags = meal_dto.tags
                    self.session.add(new_meal)
                    self.session.flush()
                    meal_id_map[meal_dto.id] = new_meal.id
                else:
                    errors.append(f"Skipped meal '{meal_dto.meal_name}': missing main recipe reference")
            restored_counts["meals"] = len(backup.data.meals)

            # 6. Restore planner entries (using mapped meal IDs)
            for pe_dto in backup.data.planner_entries:
                new_meal_id = meal_id_map.get(pe_dto.meal_id)
                if new_meal_id:
                    new_pe = PlannerEntry(
                        meal_id=new_meal_id,
                        position=pe_dto.position,
                        is_completed=pe_dto.is_completed,
                        completed_at=pe_dto.completed_at,
                        scheduled_date=pe_dto.scheduled_date,
                        shopping_mode=pe_dto.shopping_mode,
                        is_cleared=pe_dto.is_cleared,
                    )
                    self.session.add(new_pe)
                else:
                    errors.append(f"Skipped planner entry: missing meal reference")
            restored_counts["planner_entries"] = len(backup.data.planner_entries)

            # 7. Restore shopping items (only manual items - recipe items will be regenerated by sync)
            manual_items_count = 0
            for si_dto in backup.data.shopping_items:
                if si_dto.source == "manual":
                    new_si = ShoppingItem(
                        ingredient_name=si_dto.ingredient_name,
                        quantity=si_dto.quantity,
                        unit=si_dto.unit,
                        category=si_dto.category,
                        source=si_dto.source,
                        have=si_dto.have,
                        flagged=si_dto.flagged,
                    )
                    self.session.add(new_si)
                    manual_items_count += 1
            restored_counts["shopping_items"] = manual_items_count

            self.session.commit()

            # 8. Sync shopping list to regenerate recipe items from planner
            from ..services.shopping_service import ShoppingService
            shopping_service = ShoppingService(self.session)
            shopping_service.sync_shopping_list()

        except Exception as e:
            self.session.rollback()
            return RestoreResultDTO(
                success=False,
                restored_counts={},
                errors=[f"Database error: {str(e)}"],
                settings=backup.settings,
            )

        return RestoreResultDTO(
            success=len(errors) == 0,
            restored_counts=restored_counts,
            errors=errors,
            settings=backup.settings,
        )
