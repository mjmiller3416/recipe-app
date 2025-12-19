"""app/core/services/data_management_service.py

Service for data management operations (import/export recipes via xlsx).
"""

# ── Imports ─────────────────────────────────────────────────────────────────────────────────────────────────
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func
from sqlalchemy.orm import Session

from ..dtos.data_management_dtos import (   
    DuplicateAction,
    DuplicateRecipeDTO,
    DuplicateResolutionDTO,
    ExportFilterDTO,
    ImportPreviewDTO,
    ImportResultDTO,
    RecipeImportRowDTO,
    ValidationErrorDTO,
)
from ..dtos.recipe_dtos import RecipeCreateDTO, RecipeIngredientDTO, RecipeUpdateDTO
from ..models.recipe import Recipe
from ..repositories.ingredient_repo import IngredientRepo
from ..repositories.recipe_repo import RecipeRepo


# ── Constants ───────────────────────────────────────────────────────────────────────────────────────────────
RECIPE_COLUMNS = [
    "recipe_name",
    "recipe_category",
    "meal_type",
    "diet_pref",
    "total_time",
    "servings",
    "directions",
    "notes",
]

INGREDIENT_COLUMNS = [
    "recipe_name",
    "recipe_category",
    "ingredient_name",
    "ingredient_category",
    "quantity",
    "unit",
]


# ── Data Management Service ─────────────────────────────────────────────────────────────────────────────────
class DataManagementService:
    """Service layer for importing and exporting recipe data."""

    def __init__(self, session: Session):
        """Initialize the service with a database session."""
        self.session = session
        self.ingredient_repo = IngredientRepo(session)
        self.recipe_repo = RecipeRepo(session, self.ingredient_repo)

    # ── Parsing ─────────────────────────────────────────────────────────────────────────────────────────────
    def parse_xlsx(
        self, file_content: bytes
    ) -> Tuple[List[RecipeImportRowDTO], List[ValidationErrorDTO]]:
        """
        Parse xlsx file content into recipe DTOs.

        Args:
            file_content: Raw bytes of the xlsx file.

        Returns:
            Tuple of (parsed recipes, validation errors).
        """
        recipes: List[RecipeImportRowDTO] = []
        errors: List[ValidationErrorDTO] = []

        try:
            workbook = load_workbook(filename=BytesIO(file_content), read_only=True)
        except Exception as e:
            errors.append(
                ValidationErrorDTO(
                    row_number=0, field="file", message=f"Invalid xlsx file: {str(e)}"
                )
            )
            return recipes, errors

        # Parse Recipes sheet
        if "Recipes" not in workbook.sheetnames:
            errors.append(
                ValidationErrorDTO(
                    row_number=0,
                    field="sheet",
                    message="Missing required 'Recipes' sheet",
                )
            )
            return recipes, errors

        recipes_sheet = workbook["Recipes"]
        recipe_rows, recipe_errors = self._parse_recipes_sheet(recipes_sheet)
        errors.extend(recipe_errors)

        # Parse Ingredients sheet (optional)
        ingredients_by_recipe: Dict[Tuple[str, str], List[RecipeIngredientDTO]] = {}
        if "Ingredients" in workbook.sheetnames:
            ingredients_sheet = workbook["Ingredients"]
            ingredients_by_recipe, ing_errors = self._parse_ingredients_sheet(
                ingredients_sheet
            )
            errors.extend(ing_errors)

        # Combine recipes with their ingredients
        for recipe_data in recipe_rows:
            key = (
                recipe_data["recipe_name"].lower(),
                recipe_data["recipe_category"].lower(),
            )
            ingredients = ingredients_by_recipe.get(key, [])

            try:
                recipe = RecipeImportRowDTO(
                    recipe_name=recipe_data["recipe_name"],
                    recipe_category=recipe_data["recipe_category"],
                    meal_type=recipe_data.get("meal_type") or "Dinner",
                    diet_pref=recipe_data.get("diet_pref"),
                    total_time=recipe_data.get("total_time"),
                    servings=recipe_data.get("servings"),
                    directions=recipe_data.get("directions"),
                    notes=recipe_data.get("notes"),
                    ingredients=ingredients,
                )
                recipes.append(recipe)
            except Exception as e:
                errors.append(
                    ValidationErrorDTO(
                        row_number=recipe_data.get("_row", 0),
                        field="recipe",
                        message=str(e),
                    )
                )

        workbook.close()
        return recipes, errors

    def _parse_recipes_sheet(
        self, sheet: Worksheet
    ) -> Tuple[List[Dict], List[ValidationErrorDTO]]:
        """Parse the Recipes sheet into row dictionaries."""
        rows: List[Dict] = []
        errors: List[ValidationErrorDTO] = []

        # Get header row
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = [h.lower().strip() if h else "" for h in headers]

        # Validate required columns
        if "recipe_name" not in headers or "recipe_category" not in headers:
            errors.append(
                ValidationErrorDTO(
                    row_number=1,
                    field="headers",
                    message="Missing required columns: recipe_name, recipe_category",
                )
            )
            return rows, errors

        # Map column indices
        col_map = {h: i for i, h in enumerate(headers) if h}

        # Parse data rows
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            cells = [cell.value for cell in row]

            # Skip empty rows
            if not any(cells):
                continue

            recipe_name = self._get_cell_value(cells, col_map, "recipe_name")
            recipe_category = self._get_cell_value(cells, col_map, "recipe_category")

            # Validate required fields
            if not recipe_name:
                errors.append(
                    ValidationErrorDTO(
                        row_number=row_num,
                        field="recipe_name",
                        message="Recipe name is required",
                    )
                )
                continue

            if not recipe_category:
                errors.append(
                    ValidationErrorDTO(
                        row_number=row_num,
                        field="recipe_category",
                        message="Recipe category is required",
                    )
                )
                continue

            row_data = {
                "_row": row_num,
                "recipe_name": str(recipe_name).strip(),
                "recipe_category": str(recipe_category).strip(),
                "meal_type": self._get_cell_value(cells, col_map, "meal_type"),
                "diet_pref": self._get_cell_value(cells, col_map, "diet_pref"),
                "total_time": self._parse_int(
                    self._get_cell_value(cells, col_map, "total_time")
                ),
                "servings": self._parse_int(
                    self._get_cell_value(cells, col_map, "servings")
                ),
                "directions": self._get_cell_value(cells, col_map, "directions"),
                "notes": self._get_cell_value(cells, col_map, "notes"),
            }
            rows.append(row_data)

        return rows, errors

    def _parse_ingredients_sheet(
        self, sheet: Worksheet
    ) -> Tuple[Dict[Tuple[str, str], List[RecipeIngredientDTO]], List[ValidationErrorDTO]]:
        """Parse the Ingredients sheet into a dict keyed by (recipe_name, recipe_category)."""
        ingredients: Dict[Tuple[str, str], List[RecipeIngredientDTO]] = {}
        errors: List[ValidationErrorDTO] = []

        # Get header row
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        headers = [h.lower().strip() if h else "" for h in headers]

        # Validate required columns
        required = ["recipe_name", "recipe_category", "ingredient_name", "ingredient_category"]
        missing = [col for col in required if col not in headers]
        if missing:
            errors.append(
                ValidationErrorDTO(
                    row_number=1,
                    field="headers",
                    message=f"Missing required columns: {', '.join(missing)}",
                )
            )
            return ingredients, errors

        col_map = {h: i for i, h in enumerate(headers) if h}

        # Parse data rows
        for row_num, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            cells = [cell.value for cell in row]

            # Skip empty rows
            if not any(cells):
                continue

            recipe_name = self._get_cell_value(cells, col_map, "recipe_name")
            recipe_category = self._get_cell_value(cells, col_map, "recipe_category")
            ingredient_name = self._get_cell_value(cells, col_map, "ingredient_name")
            ingredient_category = self._get_cell_value(cells, col_map, "ingredient_category")

            if not all([recipe_name, recipe_category, ingredient_name, ingredient_category]):
                errors.append(
                    ValidationErrorDTO(
                        row_number=row_num,
                        field="ingredient",
                        message="Missing required ingredient fields",
                    )
                )
                continue

            recipe_key = (str(recipe_name).lower().strip(), str(recipe_category).lower().strip())
            quantity = self._parse_float(self._get_cell_value(cells, col_map, "quantity"))
            unit = self._get_cell_value(cells, col_map, "unit")

            ing_name = str(ingredient_name).strip()
            ing_cat = str(ingredient_category).strip()
            ing_unit = str(unit).strip() if unit else None

            if recipe_key not in ingredients:
                ingredients[recipe_key] = {}

            # Deduplicate by ingredient name + category (combine quantities)
            ing_key = (ing_name.lower(), ing_cat.lower())
            if ing_key in ingredients[recipe_key]:
                # Combine quantities if same unit, otherwise keep first
                existing = ingredients[recipe_key][ing_key]
                if quantity and existing.quantity and existing.unit == ing_unit:
                    existing.quantity += quantity
            else:
                ingredients[recipe_key][ing_key] = RecipeIngredientDTO(
                    ingredient_name=ing_name,
                    ingredient_category=ing_cat,
                    quantity=quantity,
                    unit=ing_unit,
                )

        # Convert dict values to lists
        result: Dict[Tuple[str, str], List[RecipeIngredientDTO]] = {}
        for recipe_key, ing_dict in ingredients.items():
            result[recipe_key] = list(ing_dict.values())

        return result, errors

    def _get_cell_value(self, cells: List, col_map: Dict[str, int], column: str):
        """Get cell value by column name, returns None if column doesn't exist."""
        if column not in col_map:
            return None
        idx = col_map[column]
        if idx >= len(cells):
            return None
        value = cells[idx]
        if isinstance(value, str):
            value = value.strip()
            return value if value else None
        return value

    def _parse_int(self, value) -> Optional[int]:
        """Parse a value to int, returns None if invalid."""
        if value is None:
            return None
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None

    def _parse_float(self, value) -> Optional[float]:
        """Parse a value to float, returns None if invalid."""
        if value is None:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    # ── Preview ─────────────────────────────────────────────────────────────────────────────────────────────
    def get_import_preview(
        self, recipes: List[RecipeImportRowDTO]
    ) -> ImportPreviewDTO:
        """
        Check parsed recipes against database to identify duplicates.

        Args:
            recipes: List of parsed recipe DTOs.

        Returns:
            ImportPreviewDTO with counts and duplicate info.
        """
        duplicates: List[DuplicateRecipeDTO] = []
        new_count = 0

        for idx, recipe in enumerate(recipes, start=2):  # Row 2 is first data row
            existing = self._find_existing_recipe(
                recipe.recipe_name, recipe.recipe_category
            )
            if existing:
                duplicates.append(
                    DuplicateRecipeDTO(
                        recipe_name=recipe.recipe_name,
                        recipe_category=recipe.recipe_category,
                        existing_id=existing.id,
                        row_number=idx,
                    )
                )
            else:
                new_count += 1

        return ImportPreviewDTO(
            total_recipes=len(recipes),
            new_recipes=new_count,
            duplicate_recipes=duplicates,
            validation_errors=[],
        )

    def _find_existing_recipe(self, name: str, category: str) -> Optional[Recipe]:
        """Find existing recipe by name and category (case-insensitive)."""
        return (
            self.session.query(Recipe)
            .filter(
                func.lower(Recipe.recipe_name) == name.strip().lower(),
                func.lower(Recipe.recipe_category) == category.strip().lower(),
            )
            .first()
        )

    # ── Execute Import ──────────────────────────────────────────────────────────────────────────────────────
    def execute_import(
        self,
        recipes: List[RecipeImportRowDTO],
        resolutions: List[DuplicateResolutionDTO],
    ) -> ImportResultDTO:
        """
        Execute the import with user-specified duplicate resolutions.

        Args:
            recipes: List of parsed recipe DTOs.
            resolutions: User's choices for handling duplicates.

        Returns:
            ImportResultDTO with counts and any errors.
        """
        # Build resolution lookup
        resolution_map: Dict[Tuple[str, str], DuplicateResolutionDTO] = {}
        for res in resolutions:
            key = (res.recipe_name.lower(), res.recipe_category.lower())
            resolution_map[key] = res

        created_count = 0
        updated_count = 0
        skipped_count = 0
        errors: List[str] = []

        for recipe in recipes:
            key = (recipe.recipe_name.lower(), recipe.recipe_category.lower())
            existing = self._find_existing_recipe(
                recipe.recipe_name, recipe.recipe_category
            )

            try:
                if existing:
                    resolution = resolution_map.get(key)
                    if resolution is None or resolution.action == DuplicateAction.SKIP:
                        skipped_count += 1
                        continue
                    elif resolution.action == DuplicateAction.UPDATE:
                        self._update_existing_recipe(existing, recipe)
                        updated_count += 1
                    elif resolution.action == DuplicateAction.RENAME:
                        if not resolution.new_name:
                            errors.append(
                                f"No new name provided for '{recipe.recipe_name}'"
                            )
                            skipped_count += 1
                            continue
                        self._create_recipe(recipe, new_name=resolution.new_name)
                        created_count += 1
                else:
                    self._create_recipe(recipe)
                    created_count += 1

            except Exception as e:
                errors.append(f"Error importing '{recipe.recipe_name}': {str(e)}")
                # Rollback to clear the failed transaction so we can continue
                self.session.rollback()

        # Commit all changes
        try:
            self.session.commit()
        except Exception as e:
            self.session.rollback()
            return ImportResultDTO(
                success=False,
                created_count=0,
                updated_count=0,
                skipped_count=len(recipes),
                errors=[f"Database error: {str(e)}"],
            )

        return ImportResultDTO(
            success=len(errors) == 0,
            created_count=created_count,
            updated_count=updated_count,
            skipped_count=skipped_count,
            errors=errors,
        )

    def _create_recipe(
        self, recipe: RecipeImportRowDTO, new_name: Optional[str] = None
    ) -> Recipe:
        """Create a new recipe from import data."""
        create_dto = RecipeCreateDTO(
            recipe_name=new_name or recipe.recipe_name,
            recipe_category=recipe.recipe_category,
            meal_type=recipe.meal_type,
            diet_pref=recipe.diet_pref,
            total_time=recipe.total_time,
            servings=recipe.servings,
            directions=recipe.directions,
            notes=recipe.notes,
            ingredients=recipe.ingredients,
        )
        return self.recipe_repo.persist_recipe_and_links(create_dto)

    def _update_existing_recipe(
        self, existing: Recipe, recipe: RecipeImportRowDTO
    ) -> Recipe:
        """Update an existing recipe with import data."""
        update_dto = RecipeUpdateDTO(
            meal_type=recipe.meal_type,
            diet_pref=recipe.diet_pref,
            total_time=recipe.total_time,
            servings=recipe.servings,
            directions=recipe.directions,
            notes=recipe.notes,
            ingredients=recipe.ingredients,
        )
        return self.recipe_repo.update_recipe(existing.id, update_dto)

    # ── Export ──────────────────────────────────────────────────────────────────────────────────────────────
    def export_recipes_to_xlsx(
        self, filter_dto: Optional[ExportFilterDTO] = None
    ) -> bytes:
        """
        Export recipes to xlsx format.

        Args:
            filter_dto: Optional filters for which recipes to export.

        Returns:
            Bytes of the xlsx file.
        """
        # Build query
        query = self.session.query(Recipe)

        if filter_dto:
            if filter_dto.recipe_category:
                query = query.filter(Recipe.recipe_category == filter_dto.recipe_category)
            if filter_dto.meal_type:
                query = query.filter(Recipe.meal_type == filter_dto.meal_type)
            if filter_dto.favorites_only:
                query = query.filter(Recipe.is_favorite == True)

        recipes = query.all()

        # Create workbook
        workbook = Workbook()

        # Recipes sheet
        recipes_sheet = workbook.active
        recipes_sheet.title = "Recipes"
        recipes_sheet.append(RECIPE_COLUMNS)

        for recipe in recipes:
            recipes_sheet.append([
                recipe.recipe_name,
                recipe.recipe_category,
                recipe.meal_type,
                recipe.diet_pref,
                recipe.total_time,
                recipe.servings,
                recipe.directions,
                recipe.notes,
            ])

        # Ingredients sheet
        ingredients_sheet = workbook.create_sheet("Ingredients")
        ingredients_sheet.append(INGREDIENT_COLUMNS)

        for recipe in recipes:
            for ing in recipe.ingredients:
                ingredients_sheet.append([
                    recipe.recipe_name,
                    recipe.recipe_category,
                    ing.ingredient.ingredient_name,
                    ing.ingredient.ingredient_category,
                    ing.quantity,
                    ing.unit,
                ])

        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()

    def generate_template_xlsx(self) -> bytes:
        """
        Generate an empty xlsx template with correct headers.

        Returns:
            Bytes of the template xlsx file.
        """
        workbook = Workbook()

        # Recipes sheet
        recipes_sheet = workbook.active
        recipes_sheet.title = "Recipes"
        recipes_sheet.append(RECIPE_COLUMNS)

        # Add example row
        recipes_sheet.append([
            "Example Recipe",
            "Main Course",
            "Dinner",
            "",
            30,
            4,
            "1. Step one\n2. Step two",
            "Optional notes here",
        ])

        # Ingredients sheet
        ingredients_sheet = workbook.create_sheet("Ingredients")
        ingredients_sheet.append(INGREDIENT_COLUMNS)

        # Add example rows
        ingredients_sheet.append([
            "Example Recipe",
            "Main Course",
            "Chicken Breast",
            "Meat",
            2,
            "lbs",
        ])
        ingredients_sheet.append([
            "Example Recipe",
            "Main Course",
            "Olive Oil",
            "Pantry",
            2,
            "tbsp",
        ])

        # Save to bytes
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.getvalue()
